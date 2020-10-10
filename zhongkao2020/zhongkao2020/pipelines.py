# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://docs.scrapy.org/en/latest/topics/item-pipeline.html


# useful for handling different item types with a single interface
# from itemadapter import ItemAdapter
import openpyxl
from collections import defaultdict
from openpyxl.chart import (
    LineChart,
    Reference,
)


class Zhongkao2020Pipeline:
    def open_spider(self, spider):
        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active
        self.summary = defaultdict(int)

    def close_spider(self, spider):
        self.sheet = self.workbook[self.workbook.sheetnames[0]]
        for key, value in self.summary.items():
            self.sheet.append([key, value])
        values = Reference(self.sheet,
                           min_row=2,
                           max_row=self.sheet.max_row,
                           min_col=2,
                           max_col=2)
        cats = Reference(self.sheet,
                         min_col=1,
                         min_row=2,
                         max_row=self.sheet.max_row,
                         )
        chart = LineChart()
        chart.style = 13
        chart.title = "成绩统计"
        chart.y_axis.title = 'amount'
        chart.x_axis.title = 'period'
        chart.add_data(values)
        chart.set_categories(cats)
        self.sheet.add_chart(chart, 'd2')
        self.workbook.save("2020score.xlsx")

    def process_item(self, item, spider):
        _title = item['name']
        _list = item['scorelist']
        self.sheet = self.workbook.create_sheet(_title)
        self.sheet.append([_title])
        for _item in _list:
            self.sheet.append(_item)
            try:
                self.summary[_item[0]] += eval(_item[1])
            except Exception:
                pass
        self.workbook.save("2020score.xlsx")
        # return item
