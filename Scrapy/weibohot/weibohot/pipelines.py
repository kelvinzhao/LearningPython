# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://docs.scrapy.org/en/latest/topics/item-pipeline.html


# useful for handling different item types with a single interface
# from itemadapter import ItemAdapter
import openpyxl
from datetime import datetime
import os

checkFile = r"isRunning"


class WeibohotPipeline:
    def open_spider(self, spider):
        f = open(checkFile, 'w')
        f.close()
        self.today = datetime.now().strftime("%d%m%Y")
        self.filename = "weibohot" + self.today + ".xlsx"
        themoment = datetime.now().strftime("%d-%m-%Y-%H-%M-%S")
        try:
            self.wb = openpyxl.load_workbook(self.filename)
        except Exception:
            self.wb = openpyxl.Workbook()
        self.sheet = self.wb.create_sheet(themoment)
        self.sheet.append(['排名', '热搜', '热度'])

    def close_spider(self, spider):
        self.wb.save(self.filename)
        isFileExist = os.path.isfile(checkFile)
        if isFileExist:
            os.remove(checkFile)

    def process_item(self, item, spider):
        self.sheet.append([item['hid'], item['hotword'], item['redu']])
        # self.wb.save(self.filename)
        return item
